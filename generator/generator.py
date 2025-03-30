import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def analyze_instance(file_path):
    """Analyze the book scanning instance file and return metrics"""
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline().strip().split()
            B = int(first_line[0])
            L = int(first_line[1])
            D = int(first_line[2])
            
            book_scores = list(map(int, f.readline().strip().split()))
            
            total_books_in_libraries = 0
            total_signup_time = 0
            total_shipping_capacity = 0
            
            for _ in range(L):
                lib_line = f.readline().strip().split()
                N_j = int(lib_line[0])
                T_j = int(lib_line[1])
                M_j = int(lib_line[2])
                
                _ = f.readline()
                
                total_books_in_libraries += N_j
                total_signup_time += T_j
                total_shipping_capacity += M_j
        
        avg_book_score = sum(book_scores) / B if B > 0 else 0
        avg_books_per_library = total_books_in_libraries / L if L > 0 else 0
        avg_signup_time = total_signup_time / L if L > 0 else 0
        avg_shipping_capacity = total_shipping_capacity / L if L > 0 else 0
        
        return {
            "No. of books": B,
            "No. of libraries": L,
            "No. of days": D,
            "Average book score": round(avg_book_score, 2),
            "Average no. of books per library": round(avg_books_per_library, 2),
            "Average sign up time": round(avg_signup_time, 2),
            "Average No. of Books Shipped per Library per Day": round(avg_shipping_capacity, 2)
        }
    
    except FileNotFoundError:
        print(f"Error: Input file not found at {file_path}")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def create_excel_report(input_file_path, output_file_path, instance_name="a_example"):
    """Create Excel report with metrics for the specified instance"""
    metrics = analyze_instance(input_file_path)
    if not metrics:
        return False
    
    data = {
        "Instance": ["UPFIEK", "a_example", "b_read_on", "c_incunabula", 
                    "e_so_many_books", "d_tough_choices", 
                    "f_libraries_of_the_world", "B5000_L90_D21", 
                    "B50000_L400_D28", "B100000_L600_D28", 
                    "B90000_L850_D21", "B95000_L2000_D28"],
        "No. of books": [""] * 12,
        "No. of libraries": [""] * 12,
        "No. of days": [""] * 12,
        "Average book score": [""] * 12,
        "Average no. of books per library": [""] * 12,
        "Average sign up time": [""] * 12,
        "Average No. of Books Shipped per Library per Day": [""] * 12
    }
    
    try:
        instance_index = data["Instance"].index(instance_name)
        for key, value in metrics.items():
            data[key][instance_index] = value
        
        df = pd.DataFrame(data)
        df.to_excel(output_file_path, index=False)
        
        wb = load_workbook(output_file_path)
        ws = wb.active
        
        header_font = Font(bold=True)
        for cell in ws[1]:  
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file_path)
        print(f"Successfully generated report: {output_file_path}")
        return True
    
    except ValueError:
        print(f"Error: Instance '{instance_name}' not found in template")
        return False
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return False

if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder = os.path.join(os.path.dirname(current_dir), "input")
    output_folder = current_dir  # Output in generator folder
    
    input_file = os.path.join(input_folder, "a_example.txt")
    output_file = os.path.join(output_folder, "library_metrics.xlsx")
    
    if not os.path.exists(input_file):
        print(f"Error: Input file not found at {input_file}")
        print("Please ensure:")
        print("1. The 'input' folder exists at the same level as 'generator'")
        print("2. The input file 'a_example.txt' exists in the input folder")
    else:
        success = create_excel_report(
            input_file_path=input_file,
            output_file_path=output_file,
            instance_name="a_example"
        )
        
        if success:
            print("Report generated successfully!")
            print(f"Location: {output_file}")
        else:
            print("Failed to generate report")
